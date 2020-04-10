import openpyxl
from jinja2 import Environment,FileSystemLoader
from jnpr.junos import Device
from jnpr.junos.utils.config import Config
from collections import defaultdict
from time import sleep
from lxml import etree

config_container=[]
device_node_ip={}

vsrx_config_excel=openpyxl.load_workbook("实验工作表.xlsx")
vsrx_int_config=vsrx_config_excel.get_sheet_by_name("接口配置")
vsrx_policy_config=vsrx_config_excel.get_sheet_by_name("安全策略")

jinja_loader=FileSystemLoader('./')
env=Environment(loader=jinja_loader)
vsrx_int_template=env.get_template('vsrx_int_config')
vsrx_sepolicy_template=env.get_template('vsrx_policy_config')

each=2
while each <= vsrx_int_config.max_row:
    device_name=vsrx_int_config["A"+str(each)].value
    device_ip=vsrx_int_config["B"+str(each)].value
    config_paramter={"int_name":vsrx_int_config["C"+str(each)].value,
                     "int_ipv4_addr": vsrx_int_config["D" + str(each)].value,
                     "area_number": vsrx_int_config["G" + str(each)].value,
                     "zone_name": vsrx_int_config["H" + str(each)].value}
    routing_config=vsrx_int_template.render(config_paramter)
    config_container.append({device_name:routing_config})
    device_node_ip[device_name]=device_ip
    each+=1

each=2
while each <= vsrx_policy_config.max_row:
    device_name=vsrx_policy_config["A"+str(each)].value
    config_paramter={"source_zone":vsrx_policy_config["C"+str(each)].value,
                     "dest_zone": vsrx_policy_config["D" + str(each)].value,
                     "source_addr": vsrx_policy_config["F" + str(each)].value,
                     "dest_addr": vsrx_policy_config["G" + str(each)].value,
                     "app": vsrx_policy_config["H" + str(each)].value,
                     "action": vsrx_policy_config["I" + str(each)].value}
    secpolicy_config=vsrx_sepolicy_template.render(config_paramter)
    config_container.append({device_name:secpolicy_config})
    each+=1
final_config=defaultdict(list)

for each in config_container:
    for key,value in each.items():
        final_config[key].append(value)

for name,ip in device_node_ip.items():
    print("#"*10)
    print("#"*10)
    print("设备:{}, ip地址:{}".format(name,ip))
    print("#"*10)
    vSRX=Device(host=ip,user="root",passwd="juniper123").open()
    vSRX_config=Config(vSRX,mode="exclusive")
    print("#"*10)
    print("以下配置将被导入进设备{}:".format(name))
    print("#"*10)
    for each in final_config[name]:
        print(each)
        vSRX_config.load(each,format="set")
    if vSRX_config.diff() == None:
        print("无新增配置，连接关闭！")
        vSRX.close()
    else:
        print("{} show | compare输出内容如下，请确认:".format(name))
        vSRX_config.pdiff()
        yes_no=input("请输入yes 或者 no 来确定是否提交上述配置：")
        while not(yes_no=="yes" or yes_no=="no")
            print("输入错误，请重新输入yes或者no！")
            yes_no=input("请输入yes 或者 no 来确定是否提交上述配置：")
        if yes_no=="yes":
            vSRX_config.commit(comment="add ospf adn security config bu PyEZ",timeout=60)
            vSRX.close()
        else:
            print("连接关闭，谢谢！")
            vSRX.close()



